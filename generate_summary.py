"""
generate_summary.py
Reads 'Salinan dari Expense kapal.xlsx' and generates SUMMARY.md
with an overall voyage-expense summary.

Requirements: pip install openpyxl
"""

import openpyxl
import os
import sys
from collections import OrderedDict

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Salinan dari Expense kapal.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "SUMMARY.md")

# Hard-coded voyage metadata extracted from the structured spreadsheet.
# Each voyage sheet uses a different layout, so we define the structure
# explicitly to ensure accuracy.
VOYAGES = OrderedDict([
    ("Voyage 1", {
        "sheet": "Voyage 1 Pengerang -> Uban -> T",
        "route": "Malaysia Loading → Indonesia Discharging → Singapore Bunkering",
        "ports": "Pengerang → Tg. Uban → Tg. Priok → Singapore",
        "duration": "July 25 – Aug 26 (32 Days)",
        "summary_idr_col": "K",  # sidebar summary IDR column
        "summary_usd_col": "L",  # sidebar summary USD column
        "opex_idr_col": "G",     # OPEX detail IDR column
        "opex_usd_col": "H",     # OPEX detail USD column
    }),
    ("Voyage 2", {
        "sheet": "Voyage 2 Singapore -> Uban -> T",
        "route": "Singapore Loading → Indonesia Discharging → Singapore Bunkering",
        "ports": "EOPL → Vopak Singapore → Tg. Uban → Tg. Priok → Singapore",
        "duration": "Sep 05 – Sep 30 (25 Days)",
        "summary_idr_col": "K",
        "summary_usd_col": "L",
        "opex_idr_col": "G",
        "opex_usd_col": "H",
    }),
    ("Voyage 3", {
        "sheet": "Voyage 3 Vopak Singapore -> Bau",
        "route": "Singapore Loading → Sulawesi Discharging → Singapore Bunkering",
        "ports": "EOPL → Vopak Singapore → Baubau → Singapore",
        "duration": "Sep 30 – Oct 30 (30 Days)",
        "summary_idr_col": "J",
        "summary_usd_col": "K",
        "opex_idr_col": "G",
        "opex_usd_col": "H",
    }),
    ("Voyage 4", {
        "sheet": "Voyage 4 Pengerang -> Baubau ->",
        "route": "Malaysia Loading → Sulawesi & East Java Discharging (Milk Run)",
        "ports": "Pengerang → Baubau → Surabaya → Pengerang",
        "duration": "Oct 30 – Nov 18 (19 Days)",
        "summary_idr_col": "K",
        "summary_usd_col": None,
        "opex_idr_col": "G",
        "opex_usd_col": "H",
    }),
    ("Voyage 5", {
        "sheet": "Voyage 5 Tg. Langsat -> Tg. Pri",
        "route": "Malaysia Loading → West Java Discharging → Singapore Bunkering",
        "ports": "Tg. Langsat → Tg. Priok → Singapore",
        "duration": "Dec 02 – Dec 15 (13 Days)",
        "summary_idr_col": "L",
        "summary_usd_col": "M",
        "opex_idr_col": "G",
        "opex_usd_col": "H",
        "opex_gbp_col": "I",
    }),
    ("Voyage 6", {
        "sheet": "Voyage 6 Singapore -> Kotabaru",
        "route": "Singapore Loading → Kalimantan Discharging → Singapore Return",
        "ports": "Singapore → Kotabaru → Singapore",
        "duration": "Dec 15 – Jan 04 (20 Days)",
        "summary_idr_col": None,
        "summary_usd_col": None,
        "opex_idr_col": "G",
        "opex_usd_col": None,
    }),
])

# Known totals from the 'Ringkasan Per Voyage' sheet (row 29).
RINGKASAN_TOTALS = {
    "Voyage 1": {"IDR": 2_009_807_145, "USD": 372_907},
    "Voyage 2": {"IDR": 213_398_392, "USD": 130_700},
    "Voyage 3": {"IDR": 45_101_860, "USD": 130_700},
    "Voyage 4": {"IDR": 513_000_518, "USD": 0},
}


def _col_letter_to_index(letter):
    """Convert column letter (e.g. 'A', 'AB') to 1-based index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _fmt_idr(value):
    """Format an IDR amount with thousand separators."""
    if value is None or value == 0:
        return "–"
    return f"Rp {value:,.0f}"


def _fmt_usd(value):
    """Format a USD amount."""
    if value is None or value == 0:
        return "–"
    return f"${value:,.2f}"


def _fmt_gbp(value):
    """Format a GBP amount."""
    if value is None or value == 0:
        return "–"
    return f"£{value:,.2f}"


def _safe_number(val):
    """Return val as a number, or 0 if it is not numeric."""
    if isinstance(val, (int, float)):
        return val
    return 0


def extract_expense_items(ws, meta):
    """Extract individual OPEX line items from a voyage sheet.

    Returns a list of dicts with keys: kategori, vendor, deskripsi, idr, usd, gbp, urgensi.
    """
    items = []
    idr_idx = _col_letter_to_index(meta["opex_idr_col"]) if meta.get("opex_idr_col") else None
    usd_idx = _col_letter_to_index(meta["opex_usd_col"]) if meta.get("opex_usd_col") else None
    gbp_idx = _col_letter_to_index(meta.get("opex_gbp_col", "") or "") if meta.get("opex_gbp_col") else None

    # Find the header row that contains "Kategori"
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
        if row and any(cell == "Kategori" for cell in row if cell):
            header_row = row_idx
            break

    if header_row is None:
        return items

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        cells = {cell.column: cell.value for cell in row}
        kategori = cells.get(1)  # Column A
        vendor = cells.get(3)    # Column C
        deskripsi = cells.get(5) # Column E
        urgensi = cells.get(6)   # Column F
        idr = _safe_number(cells.get(idr_idx)) if idr_idx else 0
        usd = _safe_number(cells.get(usd_idx)) if usd_idx else 0
        gbp = _safe_number(cells.get(gbp_idx)) if gbp_idx else 0

        if kategori and (idr or usd or gbp) and vendor:
            items.append({
                "kategori": str(kategori),
                "vendor": str(vendor),
                "deskripsi": str(deskripsi) if deskripsi else "",
                "idr": idr,
                "usd": usd,
                "gbp": gbp,
                "urgensi": str(urgensi) if urgensi else "",
            })
    return items


def compute_voyage_totals(items):
    """Sum IDR, USD, and GBP across all expense items for a voyage."""
    total_idr = sum(i["idr"] for i in items)
    total_usd = sum(i["usd"] for i in items)
    total_gbp = sum(i["gbp"] for i in items)
    return total_idr, total_usd, total_gbp


def compute_category_breakdown(items):
    """Aggregate expenses by kategori."""
    cats = OrderedDict()
    for item in items:
        k = item["kategori"]
        if k not in cats:
            cats[k] = {"idr": 0, "usd": 0, "gbp": 0}
        cats[k]["idr"] += item["idr"]
        cats[k]["usd"] += item["usd"]
        cats[k]["gbp"] += item["gbp"]
    return cats


def compute_vendor_breakdown(all_items):
    """Aggregate expenses by vendor across all voyages."""
    vendors = {}
    for item in all_items:
        v = item["vendor"]
        if v not in vendors:
            vendors[v] = {"idr": 0, "usd": 0, "gbp": 0, "count": 0}
        vendors[v]["idr"] += item["idr"]
        vendors[v]["usd"] += item["usd"]
        vendors[v]["gbp"] += item["gbp"]
        vendors[v]["count"] += 1
    # Sort by IDR descending
    return OrderedDict(sorted(vendors.items(), key=lambda x: x[1]["idr"], reverse=True))


def generate_summary():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file not found at {EXCEL_FILE}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    lines = []

    lines.append("# Voyage Expense Summary")
    lines.append("")
    lines.append("> Auto-generated from `Salinan dari Expense kapal.xlsx`")
    lines.append("> Run `python generate_summary.py` to regenerate.")
    lines.append("")

    # ── Overall fleet statistics ──────────────────────────────────────
    grand_idr = 0
    grand_usd = 0
    grand_gbp = 0
    total_days = 0
    all_items = []
    voyage_data = OrderedDict()

    for name, meta in VOYAGES.items():
        ws = wb[meta["sheet"]]
        items = extract_expense_items(ws, meta)
        t_idr, t_usd, t_gbp = compute_voyage_totals(items)
        grand_idr += t_idr
        grand_usd += t_usd
        grand_gbp += t_gbp
        all_items.extend(items)

        # Parse days from duration string, e.g. "(32 Days)"
        dur = meta["duration"]
        days = 0
        if "(" in dur and "Days)" in dur:
            days_str = dur.split("(")[1].split(" Days)")[0].strip()
            try:
                days = int(days_str)
            except ValueError:
                pass
        total_days += days

        voyage_data[name] = {
            "meta": meta,
            "items": items,
            "total_idr": t_idr,
            "total_usd": t_usd,
            "total_gbp": t_gbp,
            "days": days,
            "categories": compute_category_breakdown(items),
        }

    lines.append("## Fleet Overview")
    lines.append("")
    lines.append(f"| Metric | Value |")
    lines.append(f"|---|---|")
    lines.append(f"| Total Voyages | {len(VOYAGES)} |")
    lines.append(f"| Total Sailing Days | {total_days} |")
    lines.append(f"| Total OPEX (IDR) | {_fmt_idr(grand_idr)} |")
    lines.append(f"| Total OPEX (USD) | {_fmt_usd(grand_usd)} |")
    if grand_gbp:
        lines.append(f"| Total OPEX (GBP) | {_fmt_gbp(grand_gbp)} |")
    lines.append(f"| Total Line Items | {len(all_items)} |")
    lines.append("")

    # ── Per-Voyage Summary Table ──────────────────────────────────────
    lines.append("## Per-Voyage Summary")
    lines.append("")
    lines.append("| Voyage | Route | Duration | Days | OPEX (IDR) | OPEX (USD) | Items |")
    lines.append("|---|---|---|---:|---:|---:|---:|")
    for name, vd in voyage_data.items():
        m = vd["meta"]
        lines.append(
            f"| {name} | {m['route']} | {m['duration'].split('(')[0].strip()} "
            f"| {vd['days']} | {_fmt_idr(vd['total_idr'])} | {_fmt_usd(vd['total_usd'])} "
            f"| {len(vd['items'])} |"
        )
    lines.append(
        f"| **Grand Total** | | | **{total_days}** "
        f"| **{_fmt_idr(grand_idr)}** | **{_fmt_usd(grand_usd)}** "
        f"| **{len(all_items)}** |"
    )
    lines.append("")

    # ── Detailed per-voyage breakdown ─────────────────────────────────
    lines.append("## Voyage Details")
    lines.append("")
    for name, vd in voyage_data.items():
        m = vd["meta"]
        lines.append(f"### {name}")
        lines.append("")
        lines.append(f"- **Route:** {m['route']}")
        lines.append(f"- **Ports:** {m['ports']}")
        lines.append(f"- **Duration:** {m['duration']}")
        lines.append(f"- **Total OPEX (IDR):** {_fmt_idr(vd['total_idr'])}")
        if vd["total_usd"]:
            lines.append(f"- **Total OPEX (USD):** {_fmt_usd(vd['total_usd'])}")
        if vd["total_gbp"]:
            lines.append(f"- **Total OPEX (GBP):** {_fmt_gbp(vd['total_gbp'])}")
        lines.append("")

        if vd["categories"]:
            lines.append("| Category | IDR | USD |")
            lines.append("|---|---:|---:|")
            for cat, vals in vd["categories"].items():
                lines.append(f"| {cat} | {_fmt_idr(vals['idr'])} | {_fmt_usd(vals['usd'])} |")
            lines.append("")

        if vd["items"]:
            lines.append("<details>")
            lines.append(f"<summary>Expense line items ({len(vd['items'])} items)</summary>")
            lines.append("")
            has_gbp = any(i["gbp"] for i in vd["items"])
            if has_gbp:
                lines.append("| # | Category | Vendor | Description | Urgency | IDR | USD | GBP |")
                lines.append("|---:|---|---|---|---|---:|---:|---:|")
            else:
                lines.append("| # | Category | Vendor | Description | Urgency | IDR | USD |")
                lines.append("|---:|---|---|---|---|---:|---:|")
            for idx, item in enumerate(vd["items"], 1):
                desc = item["deskripsi"]
                if len(desc) > 80:
                    desc = desc[:77] + "..."
                row = (
                    f"| {idx} | {item['kategori']} | {item['vendor']} "
                    f"| {desc} | {item['urgensi']} "
                    f"| {_fmt_idr(item['idr'])} | {_fmt_usd(item['usd'])} "
                )
                if has_gbp:
                    row += f"| {_fmt_gbp(item['gbp'])} "
                row += "|"
                lines.append(row)
            lines.append("")
            lines.append("</details>")
            lines.append("")

    # ── Category breakdown across all voyages ─────────────────────────
    lines.append("## Expense Breakdown by Category (All Voyages)")
    lines.append("")
    all_cats = compute_category_breakdown(all_items)
    # Sort by IDR descending
    sorted_cats = sorted(all_cats.items(), key=lambda x: x[1]["idr"], reverse=True)
    lines.append("| Category | IDR | USD | % of IDR Total |")
    lines.append("|---|---:|---:|---:|")
    for cat, vals in sorted_cats:
        pct = (vals["idr"] / grand_idr * 100) if grand_idr else 0
        lines.append(f"| {cat} | {_fmt_idr(vals['idr'])} | {_fmt_usd(vals['usd'])} | {pct:.1f}% |")
    lines.append("")

    # ── Top vendors ───────────────────────────────────────────────────
    lines.append("## Top Vendors by Spend (All Voyages)")
    lines.append("")
    vendor_breakdown = compute_vendor_breakdown(all_items)
    lines.append("| Vendor | IDR | USD | # Orders |")
    lines.append("|---|---:|---:|---:|")
    for vendor, vals in list(vendor_breakdown.items())[:15]:
        lines.append(
            f"| {vendor} | {_fmt_idr(vals['idr'])} | {_fmt_usd(vals['usd'])} | {vals['count']} |"
        )
    lines.append("")

    # ── Urgency distribution ──────────────────────────────────────────
    lines.append("## Urgency Distribution")
    lines.append("")
    urgency_map = {}
    for item in all_items:
        u = item["urgensi"].upper() if item["urgensi"] else "UNSPECIFIED"
        if u not in urgency_map:
            urgency_map[u] = {"count": 0, "idr": 0}
        urgency_map[u]["count"] += 1
        urgency_map[u]["idr"] += item["idr"]
    sorted_urg = sorted(urgency_map.items(), key=lambda x: x[1]["idr"], reverse=True)
    lines.append("| Urgency | Count | Total IDR |")
    lines.append("|---|---:|---:|")
    for urg, vals in sorted_urg:
        lines.append(f"| {urg} | {vals['count']} | {_fmt_idr(vals['idr'])} |")
    lines.append("")

    # Write output
    content = "\n".join(lines) + "\n"
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"Summary written to {OUTPUT_FILE}")
    print(f"  {len(VOYAGES)} voyages, {len(all_items)} line items")
    print(f"  Grand total IDR: {_fmt_idr(grand_idr)}")
    print(f"  Grand total USD: {_fmt_usd(grand_usd)}")


if __name__ == "__main__":
    generate_summary()
